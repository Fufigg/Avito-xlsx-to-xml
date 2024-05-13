import time

import openpyxl, yattag
from openpyxl import load_workbook
from yattag import Doc, indent

wb = load_workbook("input.xlsx")
ws = wb.worksheets[0]

# Create Yattag doc, tag and text objects
doc, tag, text = Doc().tagtext()

xml_header = ''
xml_schema = ''

doc.asis(xml_header)
doc.asis(xml_schema)

with tag("Ads", formatVersion="3", target="Avito.ru", crm_version="BitrixAvitoModule"):
    for row in ws.iter_rows(min_row=2, max_row=17794, min_col=1,
                            max_col=29):  # don't forget change max_row
        row = [cell.value for cell in row]
        with tag("Ad"):
            with tag("Id"):
                text(row[0])
            with tag("ListingFee"):
                text(row[1])
            with tag("AdStatus"):
                text("Free")
            with tag("ContactMethod"):
                text(row[2])
            with tag("ContactPhone"):
                text(row[3])
            with tag("Address"):
                text(row[4])
            with tag("Category"):
                text(row[5])
            with tag("GoodsType"):
                text(row[6])
            with tag("Title"):
                doc.asis(row[7])
            with tag("Description"):
                doc.asis(row[8])
            with tag("Price"):
                text(row[9])
            with tag("Condition"):
                text(row[10])
            with tag("Images"):
                i = 11
                while i < 21:
                    if row[i] is None:
                        i = 22
                    else:
                        doc.stag('Image', url=row[i])
                        i += 1
            if row[21] is not None:
                with tag("VideoURL"):
                    doc.asis(row[21])
            with tag("AdType"):
                text(row[22])
            with tag("Availability"):
                text(row[23])
            with tag("GoodsSubType"):
                text(row[24])
            with tag("Delivery"):
                with tag("Option"):
                    text("Свой партнер СДЭК")
            if row[25] is not None:
                with tag("DateBegin"):
                    doc.asis(str(row[25]))
result = indent(
    doc.getvalue(),
    indentation='    ',
    indent_text=False
)

with open("output.xml", "w", encoding='utf-8') as f:  # give path where you want to create
    f.write(result)
